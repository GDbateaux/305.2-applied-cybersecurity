def extract_text(local_path) -> str:
    """Extract text from a local file based on its file extension."""
    local_path = str(local_path)
    try:
        if local_path.endswith((".txt", ".md", ".csv")):
            with open(local_path, "r", encoding="utf-8") as f:
                return f.read()
        elif local_path.endswith(".pdf"):
            import pdfplumber
            with pdfplumber.open(local_path) as pdf:
                return "\n".join([p.extract_text() or "" for p in pdf.pages])
        elif local_path.endswith(".docx"):
            from docx import Document
            doc = Document(local_path)
            return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        elif local_path.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(local_path, data_only=True)
            result = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                result.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        result.append(" | ".join([str(c) if c is not None else "" for c in row]))
            return "\n".join(result)
        else:
            return f"Unsupported format: {local_path}"
    except Exception as e:
        return f"Read error: {e}"