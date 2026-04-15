import os
import requests
from dotenv import load_dotenv
from langchain_core.tools import tool
from datetime import datetime


load_dotenv()
DRIVE_ID = os.getenv("KDRIVE_DRIVE_ID")
TOKEN = os.getenv("KDRIVE_TOKEN")
BASE_URL = f"https://api.infomaniak.com"
HEADERS = {"Authorization": f"Bearer {TOKEN}"}


# Helpers for kDrive interactions
def list_information_for_customers_files():
    directory_id="42"
    url = f"{BASE_URL}/3/drive/{DRIVE_ID}/files/{directory_id}/files"
    try:
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()
        data = response.json().get("data", [])
       
        files_summary = [
            {"name": f["name"], "id": f["id"], "type": f["type"], "size": f.get("size")}
            for f in data
        ]
        return files_summary
    except requests.exceptions.RequestException as e:
        return f"Error listing files: {e}"


def download_file(file_id: str):
    meta_url = f"{BASE_URL}/3/drive/{DRIVE_ID}/files/{file_id}"


    try:
        response = requests.get(meta_url, headers=HEADERS)
        response.raise_for_status()


        data = response.json().get("data", {})


        if data.get("type") == "dir":
            return "Error: Cannot download a directory."


        filename = data.get("name", f"{file_id}.bin")


    except requests.exceptions.RequestException as e:
        return f"Error retrieving file name: {e}"


    local_dir = "kdrive_cache"
    os.makedirs(local_dir, exist_ok=True)


    local_path = os.path.join(local_dir, filename)


    download_url = f"{BASE_URL}/2/drive/{DRIVE_ID}/files/{file_id}/download"


    try:
        response = requests.get(download_url, headers=HEADERS, stream=True)
        response.raise_for_status()


        with open(local_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)


        return local_path


    except requests.exceptions.RequestException as e:
        return f"Error downloading file: {e}"


def upload_message_summary_KDrive(text_content, filename="uploaded_file.txt"):
    destination_id="38"
    url = f"{BASE_URL}/3/drive/{DRIVE_ID}/upload"


    encoded_content = text_content.encode("utf-8")
    total_size = len(encoded_content)


    params = {
        "directory_id": int(destination_id),
        "file_name": filename,
        "total_size": total_size,
        "conflict": "rename"
    }


    response = requests.post(
        url,
        headers={"Authorization": HEADERS["Authorization"]},
        params=params,
        data=encoded_content
    )


    if not response.ok:
        raise Exception(f"Failed to upload file: {response.status_code} {response.text}")
    return True



@tool
def search_kdrive(query: str) -> str:
    """Search for information in kDrive's internal documents.
    List the available files, then download and read those that seem relevant.
    Use this as your first resource before consulting the Internet.
    Parameter: query (question or keywords to search for)."""
    files = list_information_for_customers_files()
    if isinstance(files, str):
        return f"kDrive error: {files}"
    if not files:
        return "No files found in kDrive."
    files_list = "\n".join([f"- ID: {f['id']} | Name: {f['name']} | Type: {f['type']}" for f in files])
    return f"Files available in kDrive:\n{files_list}\nUse read_kdrive_file(file_id=...) to read a file."



@tool
def read_kdrive_file(file_id: str) -> str:
    """Reads the contents of a kDrive file by its ID.
    Supports: .txt, .csv, .pdf, .docx, .xlsx.
    Use this after search_kdrive to read a specific file.
    Parameter: file_id (the ID returned by search_kdrive)."""
    local_path = download_file(file_id)
    if isinstance(local_path, str) and local_path.startswith("Error"):
        return f"Download error: {local_path}"
    try:
        if local_path.endswith((".txt", ".md", ".csv")):
            with open(local_path, "r", encoding="utf-8") as f:
                return f.read()
        elif local_path.endswith(".pdf"):
            import pdfplumber
            with pdfplumber.open(local_path) as pdf:
                return "\n".join([p.extract_text() or "" for p in pdf.pages])
        elif local_path.endswith(".docx"):
            from docx import Document
            doc = Document(local_path)
            return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        elif local_path.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(local_path, data_only=True)
            result = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                result.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        result.append(" | ".join([str(c) if c is not None else "" for c in row]))
            return "\n".join(result)
        else:
            return f"Unsupported format: {local_path}"
    except Exception as e:
        return f"Read error: {e}"



@tool
def summarize_and_store_feedback(content: str, author: str, project: str) -> str:
    """Summarizes customer feedback and stores it in kDrive.
    ALWAYS call this immediately when a message contains a review, complaint, or suggestion.
    Do NOT promise to save — just call this tool directly.
    Parameters: content (customer message), author (Telegram ID or name),
    project (product or topic; use 'general' if unknown)."""
    now = datetime.now()
    filename = f"feedback_{author}_{now.strftime('%Y-%m-%d_%H-%M')}.txt"
    summary = (
        f"Date: {now.strftime('%Y-%m-%d %H:%M')}\n"
        f"Author: {author}\n"
        f"Project: {project}\n\n"
        f"Content:\n{content}\n"
    )
    result = upload_message_summary_KDrive(summary, filename)
    if result:
        return f"Feedback stored in kDrive as '{filename}'."
    return f"Error storing feedback: {result}"